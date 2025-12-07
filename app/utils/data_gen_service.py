
# app/services/data_gen_service.py
import os
import pandas as pd
from datetime import datetime
import random
from typing import Optional, Dict, Any
import logging


class DataGenService:
    """Сервис для генерации тестовых данных в формате XLSX"""

    def __init__(self, config):
        self.config = config
        self.logger = logging.getLogger(__name__)
        # Путь для сохранения файлов
        self.downloads_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            'utils',
            'downloads',
            'mpstats'
        )
        os.makedirs(self.downloads_dir, exist_ok=True)

    def create_test_xlsx_file(self, category: str) -> str:
        """Создание тестового XLSX файла на основе категории"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Определяем данные на основе категории
        if "бумага" in category.lower() or "пергамент" in category.lower():
            data = self._generate_paper_data()
            category_name = "Бумага для выпечки"
        elif "форма" in category.lower() or "посуда" in category.lower():
            data = self._generate_baking_forms_data()
            category_name = "Формы для выпечки"
        elif "электроника" in category.lower():
            data = self._generate_electronics_data()
            category_name = "Электроника"
        elif "одежда" in category.lower():
            data = self._generate_clothing_data()
            category_name = "Одежда"
        else:
            data = self._generate_default_data()
            category_name = "Общие товары"

        # Создаем DataFrame
        df = pd.DataFrame(data)

        # Сортируем по частоте
        if 'Частота WB' in df.columns:
            df = df.sort_values(by='Частота WB', ascending=False)

        # Имя файла
        safe_category = category.replace('/', '_').replace('\\', '_').replace(' ', '_')
        filename = f"mpstats_{safe_category}_{timestamp}.xlsx"
        filepath = os.path.join(self.downloads_dir, filename)

        # Сохраняем в Excel с несколькими листами
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Основной лист с данными
            df.to_excel(writer, sheet_name='ag-grid', index=False)

            # Лист с статистикой
            stats_df = self._generate_stats_sheet(df, category_name)
            stats_df.to_excel(writer, sheet_name='Статистика', index=False)

            # Лист с метаданными
            meta_df = pd.DataFrame({
                'Параметр': [
                    'Категория', 'Дата сбора', 'Количество запросов',
                    'Общая частота WB', 'Общая частота Oz', 'Средняя частота'
                ],
                'Значение': [
                    category_name,
                    datetime.now().strftime("%d.%m.%Y %H:%M"),
                    len(df),
                    df['Частота WB'].sum() if 'Частота WB' in df.columns else 0,
                    df['Частота Oz'].sum() if 'Частота Oz' in df.columns else 0,
                    df['Частота WB'].mean() if 'Частота WB' in df.columns else 0
                ]
            })
            meta_df.to_excel(writer, sheet_name='Метаданные', index=False)

        # Автонастройка ширины колонок
        from openpyxl import load_workbook
        wb = load_workbook(filepath)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)

        self.logger.info(f"Тестовый файл создан: {filepath}")
        return filepath

    def _generate_paper_data(self):
        """Генерация данных для категории 'Бумага для выпечки'"""
        base_data = []

        # Основные запросы из примера
        queries = [
            ("форма для выпечки", "форма для выпечки"),
            ("пергамент для выпечки", "пергамент для выпечки"),
            ("бумага для выпечки", "бумага для выпечки"),
            ("формы для выпечки", "форма для выпечки"),
            ("бумага для аэрогриля", "бумага для аэрогриля"),
            ("пергамент для аэрогриля", "пергамент для аэрогриля"),
            ("пергамент силиконизированный", "пергамент силиконизированный"),
            ("бумага для выпечки силиконизированная", "бумага для выпечки силиконизированная"),
            ("пергаментная бумага", "пергаментная бумага"),
            ("пергамент", "пергамент"),
        ]

        priority_items = [
            "Хозяйственные товары / Бумага для выпечки",
            "Посуда и инвентарь / Формы для запекания",
            "Хозяйственные товары / Бумага пищевая",
            "Посуда и инвентарь / Контейнеры из полимеров",
        ]

        for query, cluster in queries:
            # Базовые вариации
            for i in range(random.randint(1, 3)):
                row_query = query
                if i > 0:
                    variants = [
                        f"{query} в рулоне",
                        f"{query} в листах",
                        f"{query} {random.choice(['белая', 'коричневая', 'с рисунком'])}",
                        f"{query} {random.randint(10, 50)} м",
                    ]
                    row_query = variants[min(i - 1, len(variants) - 1)]

                base_data.append({
                    'Запрос': row_query,
                    'Кластер WB': cluster,
                    'Приоритетный предмет': random.choice(priority_items),
                    'Результаты WB': random.randint(1000, 150000),
                    'Частота WB': random.randint(5000, 120000),
                    'Частота Oz': random.randint(100, 20000),
                    'Частота кластера': random.randint(10000, 300000),
                    'Частота 365': random.randint(10000, 1000000) + random.random() * 1000,
                })

        return base_data

    def _generate_baking_forms_data(self):
        """Генерация данных для категории 'Формы для выпечки'"""
        base_data = []

        queries = [
            ("форма для выпечки", "форма для выпечки"),
            ("формы для выпекания", "форма для выпечки"),
            ("силиконовая форма", "силиконовая форма"),
            ("форма для кексов", "форма для кексов"),
            ("форма для торта", "форма для торта"),
            ("форма для хлеба", "форма для хлеба"),
        ]

        priority_items = [
            "Посуда и инвентарь / Формы для запекания",
            "Посуда и инвентарь / Контейнеры из полимеров",
            "Посуда и инвентарь / Силиконовые формы",
            "Посуда и инвентарь / Формы для кексов",
        ]

        for query, cluster in queries:
            for i in range(random.randint(1, 3)):
                row_query = query
                if i > 0:
                    variants = [
                        f"{query} {random.choice(['силиконовая', 'металлическая', 'стеклянная'])}",
                        f"{query} для {random.choice(['духовки', 'мультиварки', 'микроволновки'])}",
                        f"{query} {random.choice(['круглая', 'квадратная', 'прямоугольная'])}",
                    ]
                    row_query = variants[min(i - 1, len(variants) - 1)]

                base_data.append({
                    'Запрос': row_query,
                    'Кластер WB': cluster,
                    'Приоритетный предмет': random.choice(priority_items),
                    'Результаты WB': random.randint(2000, 120000),
                    'Частота WB': random.randint(3000, 80000),
                    'Частота Oz': random.randint(200, 15000),
                    'Частота кластера': random.randint(8000, 250000),
                    'Частота 365': random.randint(8000, 800000) + random.random() * 1000,
                })

        return base_data

    def _generate_electronics_data(self):
        """Генерация данных для категории 'Электроника'"""
        base_data = []

        queries = [
            ("смартфон", "смартфон"),
            ("наушники", "наушники"),
            ("smart watch", "умные часы"),
            ("power bank", "power bank"),
            ("чехол для телефона", "чехол для телефона"),
        ]

        priority_items = [
            "Электроника / Смартфоны и телефоны",
            "Электроника / Наушники и гарнитуры",
            "Электроника / Умные часы и браслеты",
            "Электроника / Аксессуары для телефонов",
        ]

        for query, cluster in queries:
            for i in range(random.randint(1, 4)):
                row_query = query
                if i > 0:
                    variants = [
                        f"{query} {random.choice(['2024', '2025', 'новый'])}",
                        f"{query} {random.choice(['беспроводные', 'проводные', 'bluetooth'])}",
                        f"{query} {random.choice(['для андроид', 'для iphone', 'универсальный'])}",
                    ]
                    if i - 1 < len(variants):
                        row_query = variants[i - 1]

                base_data.append({
                    'Запрос': row_query,
                    'Кластер WB': cluster,
                    'Приоритетный предмет': random.choice(priority_items),
                    'Результаты WB': random.randint(5000, 200000),
                    'Частота WB': random.randint(10000, 150000),
                    'Частота Oz': random.randint(1000, 30000),
                    'Частота кластера': random.randint(20000, 500000),
                    'Частота 365': random.randint(20000, 2000000) + random.random() * 1000,
                })

        return base_data

    def _generate_clothing_data(self):
        """Генерация данных для категории 'Одежда'"""
        base_data = []

        queries = [
            ("футболка", "футболка"),
            ("джинсы", "джинсы"),
            ("куртка", "куртка"),
            ("платье", "платье"),
            ("обувь", "обувь"),
        ]

        priority_items = [
            "Одежда / Футболки и топы",
            "Одежда / Джинсы",
            "Одежда / Куртки и пальто",
            "Одежда / Платья",
            "Обувь / Кроссовки и кеды",
        ]

        for query, cluster in queries:
            for i in range(random.randint(1, 3)):
                row_query = query
                if i > 0:
                    variants = [
                        f"{query} {random.choice(['мужская', 'женская', 'детская'])}",
                        f"{query} {random.choice(['черная', 'белая', 'синяя'])}",
                        f"{query} {random.choice(['осень-зима', 'весна-лето', '2025'])}",
                    ]
                    if i - 1 < len(variants):
                        row_query = variants[i - 1]

                base_data.append({
                    'Запрос': row_query,
                    'Кластер WB': cluster,
                    'Приоритетный предмет': random.choice(priority_items),
                    'Результаты WB': random.randint(3000, 180000),
                    'Частота WB': random.randint(8000, 120000),
                    'Частота Oz': random.randint(500, 25000),
                    'Частота кластера': random.randint(15000, 400000),
                    'Частота 365': random.randint(15000, 1500000) + random.random() * 1000,
                })

        return base_data

    def _generate_default_data(self):
        """Генерация данных по умолчанию"""
        return self._generate_paper_data()[:20]  # Берем первые 20 записей

    def _generate_stats_sheet(self, df, category_name):
        """Генерация листа со статистикой"""
        stats = []

        if len(df) > 0:
            stats.append(['Категория', category_name])
            stats.append(['Всего запросов', len(df)])
            stats.append(['Уникальных кластеров', df['Кластер WB'].nunique()])

            if 'Частота WB' in df.columns:
                stats.append(['Макс. частота WB', df['Частота WB'].max()])
                stats.append(['Мин. частота WB', df['Частота WB'].min()])
                stats.append(['Средняя частота WB', int(df['Частота WB'].mean())])
                stats.append(['Общая частота WB', int(df['Частота WB'].sum())])

            if 'Частота Oz' in df.columns:
                stats.append(['Макс. частота Oz', df['Частота Oz'].max()])
                stats.append(['Общая частота Oz', int(df['Частота Oz'].sum())])

        # Топ-10 запросов
        if len(df) > 0 and 'Частота WB' in df.columns:
            stats.append(['', ''])
            stats.append(['Топ-10 запросов по частоте WB', ''])

            top_10 = df.nlargest(10, 'Частота WB')[['Запрос', 'Частота WB']]
            for idx, row in top_10.iterrows():
                stats.append([row['Запрос'], row['Частота WB']])

        return pd.DataFrame(stats, columns=['Показатель', 'Значение'])